""""exceldata

Read contents of the Excel fie used for MultiL
"""
import sys, getopt, os.path, importlib
import os, sys
import json
import copy

# Provide our standard error handling
from utils import ErrHandle

# Processing Excel
import openpyxl
from openpyxl.utils.cell import get_column_letter
from openpyxl.cell import Cell
from openpyxl import Workbook

def process_excel(oArgs):
    """Read the Excel and convert into an Object"""

    lExtracted = []
    bBack = False
    method = "iterrows"
    oErr = ErrHandle()
    oExtracted = {}
    lFeature = []
    lData = []
    lAddFeatures = [
        { "Section": "Paper Description",
          "Feature": "email_address",
          "DataType": "string",
          "WhoEnters": "-",
          "Description": "-",
          "Comment": "-",
          "values": [ None ]
        },
        { "Section": "Paper Description",
          "Feature": "verified_by_administrators",
          "DataType": "boolean",
          "WhoEnters": "-",
          "Description": "-",
          "Comment": "-",
          "values": [ True, False ]
        },
        { "Section": "Paper Description",
          "Feature": "task_detailed_other",
          "DataType": "string",
          "WhoEnters": "Author, checked and if necessary corrected by Chantal, Elise & Sharon",
          "Description": "type of task in more detail",
          "Comment": "-",
          "values": [ None ]
        },
        { "Section": "Paper Description",
          "Feature": "linguistic_property_other",
          "DataType": "string",
          "WhoEnters": "Author, checked and if necessary corrected by Chantal, Elise & Sharon",
          "Description": "type of linguistic property studied",
          "Comment": "-",
          "values": [ None ]
        },
    ]
    oAddValues = {
      "email_address": None,
      "verified_by_administrators": False,
      "task_detailed_other": None,
      "linguistic_property_other": None
    }

    try:
        # Retrieve the parameters
        flInput = oArgs.get("input")
        flOutput = oArgs.get("output")

        if flInput is None or flOutput is None:
            return False

        # Load the Excel workbook
        wb = openpyxl.load_workbook(flInput, read_only=True, data_only=True)
        sheetnames = wb.sheetnames
        ws_data = None
        ws_features = None
        for sname in sheetnames:
            if "dataset" in sname.lower():
                ws_data = wb[sname]
            elif "explanation column" in sname.lower():
                ws_features = wb[sname]
        # Do we have the right columns
        if ws_data is None or ws_features is None:
            return False

        # We have come here, so all is going well!
        # (1) Process the worksheet with the FEATURES
        # Fields: A=Section + Field name, B=Description, C=DataType, D=WhoEnters, E=MetaRegressionModifier, F=Comment
        row_num = 2
        sSection = ""
        dict_feature = {}
        while not ws_features.cell(row_num, 1).value is None:
            # Get all relevant cell values
            sFieldName = ws_features.cell(row_num, 1).value
            sDescription = ws_features.cell(row_num, 2).value
            sDataType = ws_features.cell(row_num, 3).value
            sWhoEnters = ws_features.cell(row_num, 4).value
            sMetaRegrModi = ws_features.cell(row_num, 5).value
            sComment = ws_features.cell(row_num, 6).value

            # Check if this is a section
            if sFieldName != "" and sDataType is None:
                # This just starts a new section
                sSection = sFieldName
            else:
                # Check whether this is a list of options
                arOptions = []
                if "options: " in sDataType:
                    # Retrieve the list of options
                    sOptionList = sDataType[len("options:"):].strip()
                    arOptions = [x.strip() for x in sOptionList.split(",")]
                oFeature = dict(
                    Section=sSection, Feature=sFieldName,
                    DataType=sDataType, WhoEnters=sWhoEnters,
                    Description=sDescription
                )
                if not sMetaRegrModi is None and sMetaRegrModi != "":
                    oFeature['MetaRegressionModifier'] = sMetaRegrModi
                if not sComment is None and sComment != "":
                    oFeature['Comment'] = sComment

                # Initialize a list of values for this feature
                oFeature['values'] = arOptions

                lFeature.append(oFeature)
                dict_feature[sFieldName] = oFeature

            # Go to the next row
            row_num += 1

        # Add [lFeature] with the default ones above - provided they are not yet in there
        for oFeature in lAddFeatures:
            sFieldName = oFeature['Feature']
            # Check if it really is not in the data
            if not sFieldName in dict_feature:
                lFeature.append(oFeature)
                dict_feature[sFieldName] = oFeature

        # Add the list of features to the object
        oExtracted['Features'] = lFeature

        # (2) Process the worksheet with the DATA
        if method == "iterrows":
            # Iterate speedily
            lHeader = []
            bFirst = True
            row_num = 1
            for row in ws_data.iter_rows(min_row=1, min_col=1):
                if row[0].value is None:
                    break
                oErr.Status("Reading data row {}".format(row_num))
                if bFirst:
                    # This is the first row: Expect header
                    for cell in row:
                        sValue = cell.value.strip("\t")
                        lHeader.append(sValue)
                    
                    # Reset the first flag
                    bFirst = False
                elif not row[0].value is None:
                    # This is a datarow
                    oData = {}
                    for idx, sFieldName in enumerate(lHeader):
                        cell = row[idx]
                        value = cell.value
                        # Add k/v to oData
                        oData[sFieldName] = value
                        # Possibly add value to feature value dictionary
                        if not value in dict_feature[sFieldName]['values']:
                            dict_feature[sFieldName]['values'].append(value)
                        value = cell.value
                    # Also add the default features and values
                    for k, v in oAddValues.items():
                        if not k in oData:
                            oData[k] = v

                    # Add the data in the list
                    lData.append(oData)
                row_num += 1
        else:
            row_num = 1
            # Process the first row with field names
            lHeader = []
            col_num = 1
            while not ws_data.cell(row_num, col_num).value is None:
                lHeader.append(ws_data.cell(row_num, col_num).value)
                col_num += 1

            row_num = 2
            while not ws_data.cell(row_num, 1).value is None:
                oErr.Status("Reading data row {}".format(row_num))
                # Process the values in this row
                oData = {}
                for idx, sFieldName in enumerate(lHeader):
                    col_num = idx + 1
                    # Get the value
                    value = ws_data.cell(row_num, col_num).value
                    # Add k/v to oData
                    oData[sFieldName] = value
                    # Possibly add value to feature value dictionary
                    if not value in dict_feature[sFieldName]['values']:
                        dict_feature[sFieldName]['values'].append(value)
                # Also add the default features and values
                for k, v in oAddValues.items():
                    if not k in oData:
                        oData[k] = v

                # Add the data in the list
                lData.append(oData)

                # Go to the next row
                row_num += 1

        # Walk the features once more, sorting the lists of values
        count_features = len(lFeature)
        oErr.Status("sorting values. Number of features = {}".format(count_features))
        num_feature = 0
        for k, oItem in dict_feature.items():
            num_feature += 1
            oErr.Status("Sorting feature #{}: {}".format(num_feature, k))
            try:
                oItem['values'] = sorted(oItem['values'])
            except:
                msg = oErr.get_error_message()
                oErr.Status("Key [{}]: {}".format(k, msg))
                a = 2

        # Add the list of data to the Object
        oExtracted['Dataset'] = lData

        # Save output
        with open(flOutput, "w", encoding="utf-8") as fp:
            json.dump(oExtracted, fp, indent=2)

        # Indicate all went wel
        bBack = True
    except:
        msg = oErr.get_error_message()
        oErr.DoError("process_excel")

    return bBack

def main(prgName, argv) :
    flInput = ''        # input file name: XML with author definitions
    flOutput = ''       # output file name
    lExtracted = []

    oErr = ErrHandle()
    try:
        sSyntax = prgName + ' -i <Excel file> -o <Json output file name>'
        # get all the arguments
        try:
            # Get arguments and options
            opts, args = getopt.getopt(argv, "hi:o:", ["-ifile=", "-ofile"])
        except getopt.GetoptError:
            print(sSyntax)
            sys.exit(2)
        # Walk all the arguments
        for opt, arg in opts:
            if opt in ("-h", "--help"):
                print(sSyntax)
                sys.exit(0)
            elif opt in ("-i", "--ifile"):
                flInput = arg
            elif opt in ("-o", "--ofile"):
                flOutput = arg
        # Check if all arguments are there
        if (flInput == '' or flOutput == ''):
            oErr.DoError(sSyntax)

        # Continue with the program
        oErr.Status('Input is "' + flInput + '"')
        oErr.Status('Output is "' + flOutput + '"')

        oArgs = dict(input=flInput, output=flOutput)

        # Process the Excel
        if not process_excel(oArgs):
            oErr.DoError("Could not complete reading MultiLingual results", True)


        # All went fine  
        oErr.Status("Ready")
    except:
        # act
        oErr.DoError("main")
        return False


# ----------------------------------------------------------------------------------
# Goal :  If user calls this as main, then follow up on it
# ----------------------------------------------------------------------------------
if __name__ == "__main__":
    # Call the main function with two arguments: program name + remainder
    main(sys.argv[0], sys.argv[1:])
